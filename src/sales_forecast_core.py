from __future__ import annotations

import csv
import io
import uuid
from collections import defaultdict
from dataclasses import asdict, dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Dict, Iterable, List, Tuple

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


DATE_COL = "\uC77C\uC790"
STYLE_CODE_COL = "\uC2A4\uD0C0\uC77C\uCF54\uB4DC"
QTY_COL = "\uD310\uB9E4\uC218\uB7C9"
REV_COL = "\uD310\uB9E4\uC561"
AVG_CODE = "\uD3C9\uADE0"
WEEK_COUNT = 52
SUMMARY_HEADERS = [
    "forecast_year",
    "latest_actual_date",
    "latest_actual_week_no",
    "style_code",
    "plc_code",
    "actual_ytd_qty",
    "actual_ytd_rev",
    "avg_selling_price",
    "plc_cumulative_share",
    "projected_future_qty",
    "projected_future_rev",
    "projected_full_year_qty",
    "projected_full_year_rev",
]
WEEKLY_HEADERS = [
    "forecast_year",
    "latest_actual_date",
    "style_code",
    "plc_code",
    "week_no",
    "week_start",
    "week_end",
    "row_type",
    "plc_week_share",
    "actual_qty",
    "actual_rev",
    "forecast_qty",
    "forecast_rev",
    "final_qty",
    "final_rev",
]


@dataclass
class PlcWeek:
    week_no: int
    share: float


@dataclass
class StyleForecast:
    style_code: str
    plc_code: str
    actual_ytd_qty: float
    actual_ytd_rev: float
    avg_price: float
    plc_cum_share: float
    projected_full_qty: float
    projected_full_rev: float
    projected_future_qty: float
    projected_future_rev: float


def discover_default_plc_path(downloads_dir: Path | None = None) -> Path:
    home = Path.home()
    base_dir = downloads_dir or (home / "Downloads")
    exact = base_dir / "\uC791\uB144 \uC544\uC774\uD15C PLC.csv"
    if exact.exists():
        return exact

    matches = sorted(base_dir.glob("*PLC*.csv"), key=lambda item: item.stat().st_mtime, reverse=True)
    if matches:
        return matches[0]

    raise FileNotFoundError("PLC CSV file was not found in Downloads.")


def parse_float(value: str | None) -> float:
    if value is None:
        return 0.0
    text = str(value).strip().replace(",", "")
    if not text:
        return 0.0
    return float(text)


def parse_int(value: str | None) -> int:
    return int(parse_float(value))


def parse_sales_date(value: str) -> date:
    return datetime.strptime(str(value).strip(), "%Y%m%d").date()


def iso_week_range(year: int, week_no: int) -> Tuple[str, str]:
    week_start = date.fromisocalendar(year, week_no, 1)
    week_end = date.fromisocalendar(year, week_no, 7)
    return week_start.isoformat(), week_end.isoformat()


def read_csv_rows_from_path(path: Path) -> List[dict]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        return list(csv.DictReader(handle))


def read_csv_rows_from_bytes(file_bytes: bytes) -> List[dict]:
    text = file_bytes.decode("utf-8-sig")
    return list(csv.DictReader(io.StringIO(text)))


def load_plc_patterns(plc_path: Path) -> Tuple[Dict[str, List[PlcWeek]], List[PlcWeek]]:
    patterns: Dict[str, List[PlcWeek]] = defaultdict(list)
    average_pattern: List[PlcWeek] = []

    for row in read_csv_rows_from_path(plc_path):
        item_code = (row.get("item_code") or "").strip()
        week_no = parse_int(row.get("week_no"))
        if week_no < 1 or week_no > WEEK_COUNT:
            continue

        plc_week = PlcWeek(
            week_no=week_no,
            share=parse_float(row.get("last_year_ratio_pct")) / 100.0,
        )
        if item_code == AVG_CODE:
            average_pattern.append(plc_week)
        else:
            patterns[item_code].append(plc_week)

    for code in patterns:
        patterns[code].sort(key=lambda row: row.week_no)
    average_pattern.sort(key=lambda row: row.week_no)
    return patterns, average_pattern


def normalize_sales_rows(raw_rows: Iterable[dict]) -> List[dict]:
    rows: List[dict] = []
    for row in raw_rows:
        sales_date = parse_sales_date(row[DATE_COL])
        style_code = (row.get(STYLE_CODE_COL) or "").strip()
        rows.append(
            {
                "date": sales_date,
                "year": sales_date.year,
                "week_no": sales_date.isocalendar().week,
                "style_code": style_code,
                "plc_code": style_code[2:4] if len(style_code) >= 4 else "",
                "qty": parse_float(row.get(QTY_COL)),
                "rev": parse_float(row.get(REV_COL)),
            }
        )
    return rows


def summarize_actuals(
    sales_rows: Iterable[dict], forecast_year: int
) -> Tuple[Dict[Tuple[str, int], Dict[str, float]], Dict[str, Dict[str, float]], date]:
    weekly_actuals: Dict[Tuple[str, int], Dict[str, float]] = defaultdict(lambda: {"qty": 0.0, "rev": 0.0})
    style_totals: Dict[str, Dict[str, float]] = defaultdict(lambda: {"qty": 0.0, "rev": 0.0, "plc_code": ""})
    latest_date: date | None = None

    for row in sales_rows:
        if row["year"] != forecast_year:
            continue

        style_code = row["style_code"]
        week_no = row["week_no"]
        weekly_actuals[(style_code, week_no)]["qty"] += row["qty"]
        weekly_actuals[(style_code, week_no)]["rev"] += row["rev"]
        style_totals[style_code]["qty"] += row["qty"]
        style_totals[style_code]["rev"] += row["rev"]
        style_totals[style_code]["plc_code"] = row["plc_code"]

        if latest_date is None or row["date"] > latest_date:
            latest_date = row["date"]

    if latest_date is None:
        raise ValueError(f"No sales rows were found for forecast year {forecast_year}.")

    return weekly_actuals, style_totals, latest_date


def build_style_forecasts(
    style_totals: Dict[str, Dict[str, float]],
    patterns: Dict[str, List[PlcWeek]],
    average_pattern: List[PlcWeek],
    latest_week_no: int,
) -> Tuple[List[StyleForecast], List[str]]:
    forecasts: List[StyleForecast] = []
    warnings: List[str] = []

    for style_code in sorted(style_totals):
        totals = style_totals[style_code]
        plc_code = totals["plc_code"]
        pattern = patterns.get(plc_code)

        if not pattern:
            pattern = average_pattern
            warnings.append(
                f"{style_code}: PLC item_code '{plc_code}' was not found, so the average pattern was used."
            )

        plc_cum_share = sum(row.share for row in pattern if row.week_no <= latest_week_no)
        if plc_cum_share <= 0:
            plc_cum_share = 1.0
            warnings.append(
                f"{style_code}: PLC cumulative share up to week {latest_week_no} was 0, so fallback 100% was used."
            )

        actual_qty = totals["qty"]
        actual_rev = totals["rev"]
        avg_price = actual_rev / actual_qty if actual_qty else 0.0
        projected_full_qty = actual_qty / plc_cum_share
        projected_full_rev = projected_full_qty * avg_price
        future_share = sum(row.share for row in pattern if row.week_no > latest_week_no)
        projected_future_qty = projected_full_qty * future_share
        projected_future_rev = projected_future_qty * avg_price

        forecasts.append(
            StyleForecast(
                style_code=style_code,
                plc_code=plc_code,
                actual_ytd_qty=actual_qty,
                actual_ytd_rev=actual_rev,
                avg_price=avg_price,
                plc_cum_share=plc_cum_share,
                projected_full_qty=projected_full_qty,
                projected_full_rev=projected_full_rev,
                projected_future_qty=projected_future_qty,
                projected_future_rev=projected_future_rev,
            )
        )

    return forecasts, warnings


def week_share_map(pattern: List[PlcWeek]) -> Dict[int, float]:
    return {row.week_no: row.share for row in pattern}


def round_value(value: float) -> float:
    return round(value, 2)


def build_forecast_dataset(
    sales_csv_bytes: bytes,
    plc_path: Path,
    forecast_year: int | None = None,
    source_name: str = "uploaded_sales.csv",
) -> dict:
    patterns, average_pattern = load_plc_patterns(plc_path)
    sales_rows = normalize_sales_rows(read_csv_rows_from_bytes(sales_csv_bytes))

    year = forecast_year if forecast_year is not None else max(row["year"] for row in sales_rows)
    weekly_actuals, style_totals, latest_date = summarize_actuals(sales_rows, year)
    latest_week_no = latest_date.isocalendar().week
    forecasts, warnings = build_style_forecasts(style_totals, patterns, average_pattern, latest_week_no)

    summary_rows: List[dict] = []
    weekly_rows: List[dict] = []

    for forecast in forecasts:
        pattern = patterns.get(forecast.plc_code) or average_pattern
        shares = week_share_map(pattern)

        summary_rows.append(
            {
                "forecast_year": year,
                "latest_actual_date": latest_date.isoformat(),
                "latest_actual_week_no": latest_week_no,
                "style_code": forecast.style_code,
                "plc_code": forecast.plc_code,
                "actual_ytd_qty": round_value(forecast.actual_ytd_qty),
                "actual_ytd_rev": round_value(forecast.actual_ytd_rev),
                "avg_selling_price": round_value(forecast.avg_price),
                "plc_cumulative_share": round(forecast.plc_cum_share, 6),
                "projected_future_qty": round_value(forecast.projected_future_qty),
                "projected_future_rev": round_value(forecast.projected_future_rev),
                "projected_full_year_qty": round_value(forecast.projected_full_qty),
                "projected_full_year_rev": round_value(forecast.projected_full_rev),
            }
        )

        for week_no in range(1, WEEK_COUNT + 1):
            week_start, week_end = iso_week_range(year, week_no)
            actual = weekly_actuals.get((forecast.style_code, week_no), {"qty": 0.0, "rev": 0.0})
            share = shares.get(week_no, 0.0)
            forecast_qty = 0.0
            forecast_rev = 0.0
            row_type = "actual"

            if week_no > latest_week_no:
                forecast_qty = forecast.projected_full_qty * share
                forecast_rev = forecast_qty * forecast.avg_price
                row_type = "forecast"

            weekly_rows.append(
                {
                    "forecast_year": year,
                    "latest_actual_date": latest_date.isoformat(),
                    "style_code": forecast.style_code,
                    "plc_code": forecast.plc_code,
                    "week_no": week_no,
                    "week_start": week_start,
                    "week_end": week_end,
                    "row_type": row_type,
                    "plc_week_share": round(share, 6),
                    "actual_qty": round_value(actual["qty"]),
                    "actual_rev": round_value(actual["rev"]),
                    "forecast_qty": round_value(forecast_qty),
                    "forecast_rev": round_value(forecast_rev),
                    "final_qty": round_value(actual["qty"] + forecast_qty),
                    "final_rev": round_value(actual["rev"] + forecast_rev),
                }
            )

    return {
        "result_id": uuid.uuid4().hex,
        "source_name": source_name,
        "forecast_year": year,
        "latest_actual_date": latest_date.isoformat(),
        "latest_actual_week_no": latest_week_no,
        "style_count": len(summary_rows),
        "warnings": warnings,
        "summary_rows": summary_rows,
        "weekly_rows": weekly_rows,
    }


def write_summary_csv(rows: List[dict]) -> bytes:
    return write_csv_bytes(SUMMARY_HEADERS, rows)


def write_weekly_csv(rows: List[dict]) -> bytes:
    return write_csv_bytes(WEEKLY_HEADERS, rows)


def write_csv_bytes(fieldnames: List[str], rows: List[dict]) -> bytes:
    buffer = io.StringIO()
    writer = csv.DictWriter(buffer, fieldnames=fieldnames)
    writer.writeheader()
    writer.writerows(rows)
    return ("\ufeff" + buffer.getvalue()).encode("utf-8")


def autosize_columns(sheet) -> None:
    for idx, column_cells in enumerate(sheet.columns, start=1):
        max_length = 0
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            if len(value) > max_length:
                max_length = len(value)
        sheet.column_dimensions[get_column_letter(idx)].width = min(max(max_length + 2, 12), 24)


def style_worksheet(sheet, title_fill: str) -> None:
    sheet.freeze_panes = "A2"
    header_fill = PatternFill(fill_type="solid", fgColor=title_fill)
    header_font = Font(bold=True, color="FFFFFF")
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    autosize_columns(sheet)


def build_excel_bytes(result: dict) -> bytes:
    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "Summary"
    summary_sheet.append(SUMMARY_HEADERS)
    for row in result["summary_rows"]:
        summary_sheet.append([row[key] for key in SUMMARY_HEADERS])
    style_worksheet(summary_sheet, "0F766E")

    weekly_sheet = workbook.create_sheet("WeeklyForecast")
    weekly_sheet.append(WEEKLY_HEADERS)
    for row in result["weekly_rows"]:
        weekly_sheet.append([row[key] for key in WEEKLY_HEADERS])
    style_worksheet(weekly_sheet, "1D4ED8")

    meta_sheet = workbook.create_sheet("Meta")
    meta_sheet.append(["field", "value"])
    meta_sheet.append(["source_name", result["source_name"]])
    meta_sheet.append(["forecast_year", result["forecast_year"]])
    meta_sheet.append(["latest_actual_date", result["latest_actual_date"]])
    meta_sheet.append(["latest_actual_week_no", result["latest_actual_week_no"]])
    meta_sheet.append(["style_count", result["style_count"]])
    if result["warnings"]:
        for warning in result["warnings"]:
            meta_sheet.append(["warning", warning])
    style_worksheet(meta_sheet, "7C3AED")

    binary = io.BytesIO()
    workbook.save(binary)
    return binary.getvalue()


def result_preview(result: dict) -> dict:
    totals_qty = sum(row["projected_full_year_qty"] for row in result["summary_rows"])
    totals_rev = sum(row["projected_full_year_rev"] for row in result["summary_rows"])
    return {
        "result_id": result["result_id"],
        "source_name": result["source_name"],
        "forecast_year": result["forecast_year"],
        "latest_actual_date": result["latest_actual_date"],
        "latest_actual_week_no": result["latest_actual_week_no"],
        "style_count": result["style_count"],
        "warning_count": len(result["warnings"]),
        "total_projected_full_year_qty": round_value(totals_qty),
        "total_projected_full_year_rev": round_value(totals_rev),
        "summary_rows": result["summary_rows"],
        "weekly_rows": result["weekly_rows"],
        "warnings": result["warnings"],
    }

